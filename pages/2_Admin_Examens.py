import streamlit as st
import pandas as pd
from datetime import datetime
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.database import Database
from src.fast_scheduler import FastScheduler
from src.analytics import Analytics
from src.session import require_auth, show_user_sidebar

st.set_page_config(
    page_title="Administration Examens",
    page_icon="",
    layout="wide"
)

# Authentication - only admin_examens can access
require_auth(allowed_roles=['admin_examens'])
show_user_sidebar()

@st.cache_resource
def get_database():
    return Database()

@st.cache_resource
def get_analytics(_db):
    return Analytics(_db)

def main():
    st.title(" Administration des Examens")
    st.markdown("**Génération automatique des EDT, détection des conflits et optimisation**")
    
    db = get_database()
    analytics = get_analytics(db)
    
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        " Générer EDT",
        " Détection Conflits",
        " Examens Planifiés",
        "📚 Planning par Formation",
        " Planning par Professeur"
    ])
    
    with tab1:
        st.header(" Génération Automatique des EDT")
        
        periodes = db.get_periodes_examen(actif=True)
        
        if periodes:
            periode_options = {f"{p['nom']} ({p['date_debut']} - {p['date_fin']})": p['id'] for p in periodes}
            selected = st.selectbox("Période d'examen", list(periode_options.keys()))
            periode_id = periode_options[selected]
            
            # Get annee_universitaire
            periode_info = db.execute_query(
                "SELECT annee_universitaire FROM periodes_examen WHERE id = %s", (periode_id,)
            )
            annee_univ = periode_info[0]['annee_universitaire'] if periode_info else "2024-2025"
            
            st.markdown("---")
            
            # Status
            examens_existants = db.get_examens(periode_id)
            modules = db.execute_query("SELECT COUNT(*) as count FROM modules")
            salles = db.execute_query("SELECT COUNT(*) as count FROM lieu_examen WHERE disponible = TRUE")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Examens planifiés", len(examens_existants) if examens_existants else 0)
            with col2:
                st.metric("Modules total", modules[0]['count'] if modules else 0)
            with col3:
                st.metric("Salles disponibles", salles[0]['count'] if salles else 0)
            
            st.markdown("---")
            
            col_btn1, col_btn2 = st.columns(2)
            
            with col_btn1:
                if st.button(" Générer l'EDT", type="primary", use_container_width=True):
                    with st.spinner("Génération en cours..."):
                        try:
                            scheduler = FastScheduler(db)
                            success, result = scheduler.generate_schedule(periode_id, annee_univ)
                            
                            if success:
                                st.success(" EDT généré avec succès!")
                                st.balloons()
                                
                                col_r1, col_r2, col_r3 = st.columns(3)
                                with col_r1:
                                    st.metric("Examens planifiés", result.get('scheduled', 0))
                                with col_r2:
                                    st.metric("Échecs", result.get('failed', 0))
                                with col_r3:
                                    st.metric("Durée", f"{result.get('execution_time', 0):.2f}s")
                                
                                # Show generated exams
                                st.markdown("---")
                                st.subheader(" Examens Générés")
                                
                                examens_generes = db.get_examens(periode_id)
                                if examens_generes:
                                    df = pd.DataFrame(examens_generes)
                                    st.dataframe(df, use_container_width=True, hide_index=True)
                                    
                                    # Export button
                                    csv = df.to_csv(index=False, encoding='utf-8')
                                    st.download_button(" Exporter CSV", csv, "examens_generes.csv", "text/csv")
                            else:
                                st.error(f" Erreur: {result.get('error')}")
                        except Exception as e:
                            st.error(f" Erreur: {e}")
            
            with col_btn2:
                if st.button(" Actualiser", use_container_width=True):
                    st.rerun()
            
            # Always show existing exams below
            if examens_existants:
                st.markdown("---")
                st.subheader(" Examens Actuellement Planifiés")
                df_existants = pd.DataFrame(examens_existants)
                st.dataframe(df_existants, use_container_width=True, hide_index=True)
        else:
            st.warning("Aucune période d'examen active")
    
    with tab2:
        st.header(" Détection des Conflits")
        
        conflicts = analytics.get_conflict_summary()
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Conflits Étudiants", conflicts.get('etudiants', 0), 
                     delta="OK" if conflicts.get('etudiants', 0) == 0 else "À résoudre",
                     delta_color="normal" if conflicts.get('etudiants', 0) == 0 else "inverse")
        with col2:
            st.metric("Conflits Professeurs", conflicts.get('professeurs', 0),
                     delta="OK" if conflicts.get('professeurs', 0) == 0 else "À résoudre",
                     delta_color="normal" if conflicts.get('professeurs', 0) == 0 else "inverse")
        with col3:
            st.metric("Conflits Capacité", conflicts.get('capacite', 0),
                     delta="OK" if conflicts.get('capacite', 0) == 0 else "À résoudre",
                     delta_color="normal" if conflicts.get('capacite', 0) == 0 else "inverse")
        with col4:
            st.metric("Conflits Salles", conflicts.get('salles', 0),
                     delta="OK" if conflicts.get('salles', 0) == 0 else "À résoudre",
                     delta_color="normal" if conflicts.get('salles', 0) == 0 else "inverse")
        
        st.markdown("---")
        
        # Detailed conflicts
        if conflicts.get('etudiants', 0) > 0:
            st.subheader(" Conflits Étudiants")
            conflits_etu = db.get_conflits_etudiants()
            if conflits_etu:
                st.dataframe(pd.DataFrame(conflits_etu), use_container_width=True, hide_index=True)
        
        if conflicts.get('professeurs', 0) > 0:
            st.subheader(" Conflits Professeurs")
            conflits_prof = db.get_conflits_professeurs()
            if conflits_prof:
                st.dataframe(pd.DataFrame(conflits_prof), use_container_width=True, hide_index=True)
    
    with tab3:
        st.header(" Examens Planifiés")
        
        periodes = db.get_periodes_examen(actif=True)
        
        if periodes:
            periode_options = {f"{p['nom']}": p['id'] for p in periodes}
            selected = st.selectbox("Période", list(periode_options.keys()), key="examens_periode")
            periode_id = periode_options[selected]
            
            examens = db.get_examens(periode_id)
            
            if examens:
                df = pd.DataFrame(examens)
                st.success(f" {len(examens)} examens planifiés")
                st.dataframe(df, use_container_width=True, hide_index=True)
                
                # Export
                csv = df.to_csv(index=False, encoding='utf-8')
                st.download_button(" Exporter CSV", csv, "examens.csv", "text/csv")
            else:
                st.info("Aucun examen planifié pour cette période")
        else:
            st.warning("Aucune période d'examen active")

    with tab4:
        st.header("📚 Planning par Formation")
        
        periodes = db.get_periodes_examen(actif=True)
        
        if periodes:
            col_f1, col_f2 = st.columns(2)
            
            with col_f1:
                periode_options = {f"{p['nom']}": p['id'] for p in periodes}
                selected_periode = st.selectbox("Période", list(periode_options.keys()), key="formation_periode")
                periode_id = periode_options[selected_periode]
            
            with col_f2:
                departements = db.get_departements()
                dept_options = {"Tous les départements": None}
                dept_options.update({d['nom']: d['id'] for d in departements})
                selected_dept = st.selectbox("Département", list(dept_options.keys()), key="formation_dept")
                dept_id = dept_options[selected_dept]
            
            # Get formations
            if dept_id:
                formations = db.get_formations(dept_id)
            else:
                formations = db.get_formations()
            
            if formations:
                formation_options = {f"{f['nom']} ({f['niveau']})": f['id'] for f in formations}
                selected_formation = st.selectbox("Formation", list(formation_options.keys()), key="formation_select")
                formation_id = formation_options[selected_formation]
                
                st.markdown("---")
                
                # Get planning for formation
                planning = db.get_planning_by_formation(formation_id, periode_id)
                
                if planning:
                    st.success(f" {len(planning)} examens pour cette formation")
                    
                    df = pd.DataFrame(planning)
                    df['date'] = pd.to_datetime(df['date_heure']).dt.date
                    df['heure'] = pd.to_datetime(df['date_heure']).dt.strftime('%H:%M')
                    df['date_str'] = pd.to_datetime(df['date_heure']).dt.strftime('%d/%m/%Y')
                    
                    # Stats
                    col_s1, col_s2, col_s3 = st.columns(3)
                    with col_s1:
                        st.metric("Total examens", len(planning))
                    with col_s2:
                        st.metric("Jours d'examens", df['date'].nunique())
                    with col_s3:
                        if 'duree_minutes' in df.columns:
                            total_hours = df['duree_minutes'].sum() // 60
                            st.metric("Durée totale", f"{total_hours}h")
                    
                    st.markdown("---")
                    
                    # Display mode
                    view_mode = st.radio(
                        "Mode d'affichage",
                        [" Calendrier", " Liste Détaillée", " Tableau"],
                        horizontal=True,
                        key="view_mode_formation"
                    )
                    
                    if view_mode == " Calendrier":
                        # Group by date - Calendar view
                        dates = sorted(df['date'].unique())
                        
                        for date in dates:
                            exams_on_date = df[df['date'] == date].sort_values('heure')
                            
                            st.markdown(f"###  {date.strftime('%A %d %B %Y')}")
                            
                            for idx, exam in exams_on_date.iterrows():
                                with st.container():
                                    col1, col2, col3 = st.columns([1, 2, 1])
                                    
                                    with col1:
                                        st.markdown(f"** {exam['heure']}**")
                                        st.caption(f"{exam.get('duree_minutes', 'N/A')} min")
                                    
                                    with col2:
                                        st.markdown(f"**📖 {exam.get('module_nom', 'N/A')}**")
                                        st.caption(f"Code: {exam.get('module_code', 'N/A')}")
                                    
                                    with col3:
                                        st.markdown(f"** {exam.get('salle_nom', 'N/A')}**")
                                        st.caption(f" {exam.get('nb_inscrits', 'N/A')} étudiants")
                                
                                st.markdown("---")
                    
                    elif view_mode == " Liste Détaillée":
                        for idx, exam in df.iterrows():
                            with st.expander(f"📖 {exam.get('module_nom', 'N/A')} - {exam['date_str']} à {exam['heure']}"):
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.markdown("**📖 Informations Module**")
                                    st.write(f"- **Code:** {exam.get('module_code', 'N/A')}")
                                    st.write(f"- **Durée examen:** {exam.get('duree_minutes', 'N/A')} minutes")
                                    st.write(f"- **Inscrits:** {exam.get('nb_inscrits', 'N/A')} étudiants")
                                
                                with col2:
                                    st.markdown("** Logistique**")
                                    st.write(f"- **Date:** {exam['date_str']}")
                                    st.write(f"- **Heure:** {exam['heure']}")
                                    st.write(f"- **Salle:** {exam.get('salle_nom', 'N/A')}")
                    
                    else:  # Tableau
                        st.dataframe(df, use_container_width=True, hide_index=True)
                    
                    # Export options
                    st.markdown("---")
                    col_exp1, col_exp2 = st.columns(2)
                    
                    with col_exp1:
                        csv = df.to_csv(index=False, encoding='utf-8')
                        st.download_button(" Exporter CSV", csv, f"planning_formation_{formation_id}.csv", "text/csv")
                    
                    with col_exp2:
                        # Calendar Excel export
                        import io
                        try:
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            from openpyxl.utils import get_column_letter
                            
                            output = io.BytesIO()
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Planning Calendrier"
                            
                            dates = sorted(df['date'].unique())
                            time_slots = sorted(df['heure'].unique())
                            
                            # Styles
                            header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
                            header_font = Font(bold=True, color="FFFFFF", size=12)
                            time_fill = PatternFill(start_color="AEC7E8", end_color="AEC7E8", fill_type="solid")
                            exam_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                            border = Border(
                                left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin')
                            )
                            
                            # Title
                            ws.merge_cells('A1:' + get_column_letter(len(dates) + 1) + '1')
                            ws['A1'] = f"Planning Formation - {selected_formation}"
                            ws['A1'].font = Font(bold=True, size=14)
                            ws['A1'].alignment = Alignment(horizontal='center')
                            
                            # Headers - Days
                            ws['A3'] = "Heure"
                            ws['A3'].fill = header_fill
                            ws['A3'].font = header_font
                            ws['A3'].border = border
                            
                            for col_idx, date in enumerate(dates, start=2):
                                cell = ws.cell(row=3, column=col_idx)
                                cell.value = date.strftime('%d/%m/%Y\n%A')
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                                cell.border = border
                                ws.column_dimensions[get_column_letter(col_idx)].width = 30
                            
                            ws.column_dimensions['A'].width = 10
                            
                            # Time slots
                            for row_idx, time_slot in enumerate(time_slots, start=4):
                                time_cell = ws.cell(row=row_idx, column=1)
                                time_cell.value = time_slot
                                time_cell.fill = time_fill
                                time_cell.font = Font(bold=True)
                                time_cell.border = border
                                ws.row_dimensions[row_idx].height = 50
                                
                                for col_idx, date in enumerate(dates, start=2):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    exam = df[(df['date'] == date) & (df['heure'] == time_slot)]
                                    
                                    if not exam.empty:
                                        e = exam.iloc[0]
                                        cell.value = f"{e.get('module_code', '')}\n{e.get('module_nom', '')}\nSalle: {e.get('salle_nom', '')}\n{e.get('nb_inscrits', '')} etudiants"
                                        cell.fill = exam_fill
                                    
                                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                                    cell.border = border
                            
                            wb.save(output)
                            output.seek(0)
                            
                            st.download_button(
                                " Exporter Excel (Calendrier)",
                                output,
                                f"planning_calendrier_formation_{formation_id}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        except ImportError:
                            st.warning("openpyxl non installé pour l'export Excel")
                else:
                    st.info("Aucun examen planifié pour cette formation")
            else:
                st.warning("Aucune formation disponible")
        else:
            st.warning("Aucune période d'examen active")
    
    with tab5:
        st.header(" Planning par Professeur")
        
        periodes = db.get_periodes_examen(actif=True)
        
        if periodes:
            col_p1, col_p2 = st.columns(2)
            
            with col_p1:
                periode_options = {f"{p['nom']}": p['id'] for p in periodes}
                selected_periode = st.selectbox("Période", list(periode_options.keys()), key="prof_periode")
                periode_id = periode_options[selected_periode]
            
            with col_p2:
                departements = db.get_departements()
                dept_options = {"Tous les départements": None}
                dept_options.update({d['nom']: d['id'] for d in departements})
                selected_dept = st.selectbox("Département", list(dept_options.keys()), key="prof_dept")
                dept_id = dept_options[selected_dept]
            
            # Get professors
            if dept_id:
                professeurs = db.get_professeurs(dept_id)
            else:
                professeurs = db.get_professeurs()
            
            if professeurs:
                prof_options = {f"{p['nom']} {p['prenom']} ({p.get('grade', 'N/A')})": p['id'] for p in professeurs}
                selected_prof = st.selectbox("Professeur", list(prof_options.keys()), key="prof_select")
                prof_id = prof_options[selected_prof]
                
                st.markdown("---")
                
                # Get planning for professor
                planning = db.get_planning_professeur(prof_id, periode_id)
                
                if planning:
                    st.success(f" {len(planning)} surveillance(s) pour ce professeur")
                    
                    df = pd.DataFrame(planning)
                    df['date'] = pd.to_datetime(df['date_heure']).dt.date
                    df['heure'] = pd.to_datetime(df['date_heure']).dt.strftime('%H:%M')
                    df['date_str'] = pd.to_datetime(df['date_heure']).dt.strftime('%d/%m/%Y')
                    
                    # Stats
                    col_s1, col_s2, col_s3, col_s4 = st.columns(4)
                    with col_s1:
                        st.metric("Total surveillances", len(planning))
                    with col_s2:
                        if 'role' in df.columns:
                            responsable = len(df[df['role'] == 'responsable'])
                            st.metric("En tant que responsable", responsable)
                    with col_s3:
                        st.metric("Jours mobilisés", df['date'].nunique())
                    with col_s4:
                        if 'duree_minutes' in df.columns:
                            total_hours = df['duree_minutes'].sum() // 60
                            st.metric("Heures totales", f"{total_hours}h")
                    
                    st.markdown("---")
                    
                    # Display mode
                    view_mode = st.radio(
                        "Mode d'affichage",
                        [" Calendrier", " Liste Détaillée", " Tableau"],
                        horizontal=True,
                        key="view_mode_prof"
                    )
                    
                    if view_mode == " Calendrier":
                        # Group by date - Calendar view
                        dates = sorted(df['date'].unique())
                        
                        for date in dates:
                            exams_on_date = df[df['date'] == date].sort_values('heure')
                            
                            st.markdown(f"###  {date.strftime('%A %d %B %Y')}")
                            
                            for idx, exam in exams_on_date.iterrows():
                                with st.container():
                                    col1, col2, col3 = st.columns([1, 2, 1])
                                    
                                    with col1:
                                        st.markdown(f"** {exam['heure']}**")
                                        st.caption(f"{exam.get('duree_minutes', 'N/A')} min")
                                    
                                    with col2:
                                        st.markdown(f"**📖 {exam.get('module_nom', 'N/A')}**")
                                        role_badge = " Responsable" if exam.get('role') == 'responsable' else " Surveillant"
                                        st.caption(f"{role_badge}")
                                    
                                    with col3:
                                        st.markdown(f"** {exam.get('salle_nom', 'N/A')}**")
                                        st.caption(f" {exam.get('nb_inscrits', 'N/A')} étudiants")
                                
                                st.markdown("---")
                    
                    elif view_mode == " Liste Détaillée":
                        for idx, exam in df.iterrows():
                            role_icon = "" if exam.get('role') == 'responsable' else ""
                            with st.expander(f"{role_icon} {exam.get('module_nom', 'N/A')} - {exam['date_str']} à {exam['heure']}"):
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.markdown("**📖 Informations Examen**")
                                    st.write(f"- **Module:** {exam.get('module_nom', 'N/A')}")
                                    st.write(f"- **Rôle:** {exam.get('role', 'N/A').title()}")
                                    st.write(f"- **Durée:** {exam.get('duree_minutes', 'N/A')} minutes")
                                    st.write(f"- **Inscrits:** {exam.get('nb_inscrits', 'N/A')} étudiants")
                                
                                with col2:
                                    st.markdown("** Logistique**")
                                    st.write(f"- **Date:** {exam['date_str']}")
                                    st.write(f"- **Heure:** {exam['heure']}")
                                    st.write(f"- **Salle:** {exam.get('salle_nom', 'N/A')}")
                    
                    else:  # Tableau
                        st.dataframe(df, use_container_width=True, hide_index=True)
                    
                    # Export options
                    st.markdown("---")
                    col_exp1, col_exp2 = st.columns(2)
                    
                    with col_exp1:
                        csv = df.to_csv(index=False, encoding='utf-8')
                        st.download_button(" Exporter CSV", csv, f"planning_prof_{prof_id}.csv", "text/csv")
                    
                    with col_exp2:
                        # Calendar Excel export
                        import io
                        try:
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            from openpyxl.utils import get_column_letter
                            
                            output = io.BytesIO()
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Planning Calendrier"
                            
                            dates = sorted(df['date'].unique())
                            time_slots = sorted(df['heure'].unique())
                            
                            # Styles
                            header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
                            header_font = Font(bold=True, color="FFFFFF", size=12)
                            time_fill = PatternFill(start_color="AEC7E8", end_color="AEC7E8", fill_type="solid")
                            exam_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                            border = Border(
                                left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin')
                            )
                            
                            # Title
                            ws.merge_cells('A1:' + get_column_letter(len(dates) + 1) + '1')
                            ws['A1'] = f"Planning Professeur - {selected_prof}"
                            ws['A1'].font = Font(bold=True, size=14)
                            ws['A1'].alignment = Alignment(horizontal='center')
                            
                            # Headers - Days
                            ws['A3'] = "Heure"
                            ws['A3'].fill = header_fill
                            ws['A3'].font = header_font
                            ws['A3'].border = border
                            
                            for col_idx, date in enumerate(dates, start=2):
                                cell = ws.cell(row=3, column=col_idx)
                                cell.value = date.strftime('%d/%m/%Y\n%A')
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                                cell.border = border
                                ws.column_dimensions[get_column_letter(col_idx)].width = 30
                            
                            ws.column_dimensions['A'].width = 10
                            
                            # Time slots
                            for row_idx, time_slot in enumerate(time_slots, start=4):
                                time_cell = ws.cell(row=row_idx, column=1)
                                time_cell.value = time_slot
                                time_cell.fill = time_fill
                                time_cell.font = Font(bold=True)
                                time_cell.border = border
                                ws.row_dimensions[row_idx].height = 50
                                
                                for col_idx, date in enumerate(dates, start=2):
                                    cell = ws.cell(row=row_idx, column=col_idx)
                                    exam = df[(df['date'] == date) & (df['heure'] == time_slot)]
                                    
                                    if not exam.empty:
                                        e = exam.iloc[0]
                                        role = "Responsable" if e.get('role') == 'responsable' else "Surveillant"
                                        cell.value = f"{e.get('module_nom', '')}\n{role}\nSalle: {e.get('salle_nom', '')}\n{e.get('nb_inscrits', '')} etudiants"
                                        cell.fill = exam_fill
                                    
                                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                                    cell.border = border
                            
                            wb.save(output)
                            output.seek(0)
                            
                            st.download_button(
                                " Exporter Excel (Calendrier)",
                                output,
                                f"planning_calendrier_prof_{prof_id}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        except ImportError:
                            st.warning("openpyxl non installé pour l'export Excel")
                else:
                    st.info("Aucune surveillance planifiée pour ce professeur")
            else:
                st.warning("Aucun professeur disponible")
        else:
            st.warning("Aucune période d'examen active")

if __name__ == "__main__":
    main()
